import re
import unicodedata
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods

from cardapio.models import Cardapio, Categoria, FaixaPreco, Produto

COLUNAS_ESPERADAS = [
    "Categoria",
    "Tipo da Categoria",
    "Produto",
    "Descrição",
    "Modo de Venda",
    "Preço",
    "Preço por KG",
    "Horário Específico",
    "Ativo",
]


def _norm(valor) -> str:
    """Normaliza texto para casar valores: tira acento, maiúsculas, ignora o que vem após '('."""
    s = str(valor or "").strip()
    if "(" in s:
        s = s.split("(")[0]
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split()).upper()


def _mapa_tipo_categoria() -> dict:
    """Aceita tanto o código (PROTEINA) quanto o rótulo amigável (Proteína)."""
    m = {}
    for code, label in Categoria.Tipo.choices:
        m[_norm(code)] = code
        m[_norm(label)] = code
    return m


def _mapa_modo_venda() -> dict:
    m = {}
    for code, label in Produto.ModoVenda.choices:
        m[_norm(code)] = code
        m[_norm(label)] = code
    return m


def _preco_dec(valor) -> Decimal:
    """Converte o preço da planilha em Decimal, tolerante a 'R$', milhar e vírgula.

    Aceita: 6 · 6,00 · 159,90 · "R$ 15,50" · 1.234,56 · 159.9 (float do Excel).
    """
    if valor is None or valor == "":
        return Decimal("0.00")
    if isinstance(valor, (int, float)):
        try:
            return Decimal(str(valor))
        except InvalidOperation:
            return Decimal("0.00")
    s = str(valor).strip()
    # mantém só dígitos, vírgula, ponto e sinal
    s = re.sub(r"[^0-9,.\-]", "", s)
    if not s:
        return Decimal("0.00")
    # formato brasileiro "1.234,56" -> tira milhar, vírgula vira ponto
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0.00")


def _cel(row, idx):
    """Lê a célula com segurança (índice pode faltar ou linha ser curta)."""
    if idx < 0 or len(row) <= idx:
        return None
    return row[idx]


def _ler_linhas(arquivo):
    """Lê a planilha (.xlsx OU .csv) e devolve lista de linhas (listas de strings)."""
    nome = (arquivo.name or "").lower()
    if nome.endswith(".csv"):
        import csv
        import io
        raw = arquivo.read()
        if isinstance(raw, bytes):
            texto = None
            for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    texto = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if texto is None:
                texto = raw.decode("utf-8", errors="replace")
        else:
            texto = raw
        amostra = texto[:4000]
        # detecta o separador: ';' (comum no Excel-BR) ou ','
        delim = ";" if amostra.count(";") > amostra.count(",") else ","
        linhas = []
        for row in csv.reader(io.StringIO(texto), delimiter=delim):
            linhas.append([("" if c is None else str(c)).strip() for c in row])
        return linhas
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active
    linhas = []
    for row in ws.iter_rows(values_only=True):
        linhas.append([("" if c is None else str(c)).strip() for c in row])
    return linhas


def _e_formato_template(rows) -> bool:
    """True se a planilha é o NOSSO modelo (tem cabeçalho com Produto e Categoria)."""
    if not rows:
        return False
    header = [_norm(c) for c in rows[0]]
    return "PRODUTO" in header and "CATEGORIA" in header


# Palavras-chave p/ adivinhar o tipo de uma categoria a partir do nome (formato livre).
_TIPO_PALAVRAS = [
    ("PROTE", "PROTEINA"),
    ("ACOMP", "ACOMPANHAMENTO"), ("GUARNI", "ACOMPANHAMENTO"),
    ("SANDU", "SANDUICHE"),
    ("SOPA", "SOPA"), ("CALDO", "SOPA"),
    ("BEBIDA", "BEBIDA"), ("SUCO", "BEBIDA"), ("REFRI", "BEBIDA"), ("DRINK", "BEBIDA"),
    ("SOBREMESA", "SOBREMESA"), ("DOCE", "SOBREMESA"),
    ("ESPET", "ESPETINHO"),
    ("GRELHAD", "GRELHADO"),
    ("ADICIO", "ADICIONAL"),
]


def _tipo_por_nome(nome: str) -> str:
    n = _norm(nome)
    for chave, cod in _TIPO_PALAVRAS:
        if chave in n:
            return cod
    return "OUTRO"


def _criar_cardapio_do_form(request, nome_cardapio, exclusivo):
    """Cria o Cardápio + agenda a partir do formulário. Retorna (cardapio, tipo, qtd_dias)."""
    tipo_cardapio = request.POST.get("tipo_cardapio", "NORMAL")
    cardapio_obj = Cardapio.objects.create(
        nome=nome_cardapio, tipo=tipo_cardapio, exclusivo=exclusivo, ativo=True,
    )
    qtd_dias = 0
    if tipo_cardapio == "ESPECIAL":
        d_inicio = request.POST.get("data_inicio")
        d_fim = request.POST.get("data_fim")
        if d_inicio:
            cardapio_obj.data_inicio = d_inicio
        if d_fim:
            cardapio_obj.data_fim = d_fim
        cardapio_obj.save()
    else:
        from cardapio.models import DisponibilidadeCardapio
        periodo = request.POST.get("periodo", "ALMOCO")
        for dia in request.POST.getlist("dias"):
            if periodo == "DIA_INTEIRO":
                DisponibilidadeCardapio.objects.create(cardapio=cardapio_obj, dia_semana=int(dia), periodo="ALMOCO")
                DisponibilidadeCardapio.objects.create(cardapio=cardapio_obj, dia_semana=int(dia), periodo="JANTAR")
            else:
                DisponibilidadeCardapio.objects.create(cardapio=cardapio_obj, dia_semana=int(dia), periodo=periodo)
            qtd_dias += 1
    return cardapio_obj, tipo_cardapio, qtd_dias


def _importar_template(rows, cardapio_obj):
    """Importa o NOSSO modelo (linhas com cabeçalho Categoria/Produto/...). Retorna (criados, existentes)."""
    from datetime import datetime

    header = rows[0]

    def hidx(nome):
        alvo = _norm(nome)
        for i, h in enumerate(header):
            if _norm(h) == alvo:
                return i
        return -1

    idx_cat = hidx("Categoria")
    idx_tipo = hidx("Tipo da Categoria")
    idx_prod = hidx("Produto")
    idx_desc = hidx("Descrição")
    idx_modo = hidx("Modo de Venda")
    idx_preco = hidx("Preço")
    idx_kg = hidx("Preço por KG")
    idx_ativo = hidx("Ativo")
    idx_horario = hidx("Horário Específico")

    mapa_tipo = _mapa_tipo_categoria()
    mapa_modo = _mapa_modo_venda()
    criados = existentes = 0

    for row in rows[1:]:
        raw_prod = str(_cel(row, idx_prod)) if _cel(row, idx_prod) else ""
        raw_cat = str(_cel(row, idx_cat)) if _cel(row, idx_cat) else ""
        nome_prod = re.sub(r"\s+", " ", raw_prod.strip())
        nome_cat = re.sub(r"\s+", " ", raw_cat.strip())
        if not nome_prod or not nome_cat or nome_prod == "None" or nome_cat == "None":
            continue

        tipo_cat = "OUTRO"
        if _cel(row, idx_tipo):
            tipo_cat = mapa_tipo.get(_norm(_cel(row, idx_tipo)), "OUTRO")
        categoria = Categoria.objects.filter(nome__iexact=nome_cat).first()
        if not categoria:
            categoria = Categoria.objects.create(nome=nome_cat, tipo=tipo_cat)

        h_inicio = h_fim = None
        if _cel(row, idx_horario):
            horario_str = str(_cel(row, idx_horario)).strip()
            if "-" in horario_str:
                partes = horario_str.split("-")
                try:
                    h_inicio = datetime.strptime(partes[0].strip(), "%H:%M").time()
                    h_fim = datetime.strptime(partes[1].strip(), "%H:%M").time()
                except ValueError:
                    pass

        produto = Produto.objects.filter(nome__iexact=nome_prod).first()
        if produto:
            if h_inicio and h_fim:
                produto.horario_inicio = h_inicio
                produto.horario_fim = h_fim
                produto.save(update_fields=["horario_inicio", "horario_fim"])
            existentes += 1
        else:
            desc = ""
            if _cel(row, idx_desc) and _cel(row, idx_desc) != "None":
                desc = str(_cel(row, idx_desc)).strip()
            modo = "UNIDADE"
            if _cel(row, idx_modo) and _cel(row, idx_modo) != "None":
                modo = mapa_modo.get(_norm(_cel(row, idx_modo)), "UNIDADE")
            ativo = True
            if _cel(row, idx_ativo) and _norm(_cel(row, idx_ativo)) in ("NAO", "NO", "FALSE", "0", "N"):
                ativo = False
            produto = Produto.objects.create(
                nome=nome_prod, categoria=categoria, descricao=desc, modo_venda=modo,
                preco=_preco_dec(_cel(row, idx_preco)), preco_kg=_preco_dec(_cel(row, idx_kg)),
                ativo=ativo, horario_inicio=h_inicio, horario_fim=h_fim,
            )
            criados += 1
        cardapio_obj.produtos.add(produto)

    return criados, existentes


def _titulo_categoria(nome: str) -> str:
    """Deixa títulos de seção em MAIÚSCULA num formato bonito (REFEIÇÕES -> Refeições)."""
    n = (nome or "").strip()
    return n.title() if n and n == n.upper() else n


def _importar_livre(rows, cardapio_obj):
    """Importa o formato do CLIENTE (3 colunas, sem cabeçalho):
    títulos de seção (col A com B e C vazias) viram Categorias; os itens seguintes
    entram nessa categoria. Nome repetido com tamanhos diferentes vira produto FAIXA.

    A planilha do cliente é a FONTE DA VERDADE: casa por (nome + categoria) e
    ATUALIZA preço/modo/faixas; só cria se não existir naquela categoria (não
    reaproveita um produto de mesmo nome de OUTRA categoria).
    Retorna (criados, atualizados)."""
    from collections import OrderedDict

    cat_atual = None
    grupos = OrderedDict()
    for row in rows:
        a = re.sub(r"\s+", " ", (_cel(row, 0) or "").strip())
        b = (_cel(row, 1) or "").strip()
        c = (_cel(row, 2) or "").strip()
        if not a:
            continue
        # Título de seção = coluna A preenchida e SEM descrição e SEM preço.
        if not b and not c:
            cat_atual = a
            continue
        if not cat_atual:
            cat_atual = "Cardápio"
        key = (cat_atual.lower(), a.lower())
        grupos.setdefault(key, {"cat": cat_atual, "nome": a, "linhas": []})["linhas"].append((b, c))

    criados = atualizados = 0
    cache_cat = {}
    for g in grupos.values():
        cat_nome = _titulo_categoria(g["cat"])
        chave_cat = cat_nome.lower()
        categoria = cache_cat.get(chave_cat)
        if categoria is None:
            categoria = Categoria.objects.filter(nome__iexact=cat_nome).first()
            if not categoria:
                categoria = Categoria.objects.create(nome=cat_nome[:80], tipo=_tipo_por_nome(cat_nome))
            cache_cat[chave_cat] = categoria

        nome = g["nome"]
        linhas = g["linhas"]
        faixa = len(linhas) > 1
        if faixa:
            modo, preco, desc = "FAIXA", Decimal("0.00"), ""
        else:
            rot, pr = linhas[0]
            modo = "ADICIONAL" if "ADICIO" in _norm(nome) else "UNIDADE"
            preco, desc = _preco_dec(pr), (rot or "").strip()

        produto = Produto.objects.filter(nome__iexact=nome, categoria=categoria).first()
        if produto:
            produto.modo_venda = modo
            produto.preco = preco
            if desc:
                produto.descricao = desc
            produto.ativo = True
            produto.save()
            atualizados += 1
        else:
            produto = Produto.objects.create(
                nome=nome[:120], categoria=categoria, descricao=desc, modo_venda=modo, preco=preco,
            )
            criados += 1

        # Faixas: substitui pelas da planilha (quando o item tem vários tamanhos).
        if faixa:
            produto.faixas.all().delete()
            for i, (rot, pr) in enumerate(linhas):
                rotulo = (rot or f"Opção {i + 1}").strip()[:30]
                FaixaPreco.objects.create(produto=produto, rotulo=rotulo, preco=_preco_dec(pr), ordem=i)
        else:
            # deixou de ser faixa: limpa faixas antigas
            produto.faixas.all().delete()

        cardapio_obj.produtos.add(produto)

    return criados, atualizados


def baixar_planilha_exemplo(request):
    """Modelo no FORMATO DO CLIENTE: seções em MAIÚSCULA + (item, tamanho, preço).

    É o mesmo formato que a importação lê (round-trip). Vem pré-preenchido com o
    cardápio real do lojista, agrupado por categoria.
    """
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cardápio"
    for letra, larg in zip("ABC", [36, 26, 14]):
        ws.column_dimensions[letra].width = larg

    hdr_fill = PatternFill("solid", fgColor="D97706")
    hdr_font = Font(bold=True, color="FFFFFF", size=12)

    def _preco_txt(v):
        v = v or Decimal("0.00")
        return f"R$ {v:.2f}".replace(".", ",")

    estado = {"linha": 1}

    def secao(titulo):
        r = estado["linha"]
        ws.cell(row=r, column=1, value=(titulo or "").upper())
        ws.cell(row=r, column=2, value="")
        ws.cell(row=r, column=3, value="")
        for col in (1, 2, 3):
            c = ws.cell(row=r, column=col)
            c.fill = hdr_fill
            c.font = hdr_font
        estado["linha"] = r + 1

    def item(nome, tamanho, preco):
        r = estado["linha"]
        ws.cell(row=r, column=1, value=nome)
        ws.cell(row=r, column=2, value=tamanho or "")
        ws.cell(row=r, column=3, value=_preco_txt(preco))
        estado["linha"] = r + 1

    # Pré-preenche com o cardápio real, agrupado por categoria.
    algum = False
    for cat in Categoria.objects.prefetch_related("produtos__faixas").order_by("ordem", "nome"):
        prods = list(cat.produtos.all())
        if not prods:
            continue
        algum = True
        secao(cat.nome)
        for p in prods:
            faixas = list(p.faixas.all())
            if p.modo_venda == Produto.ModoVenda.FAIXA and faixas:
                for f in faixas:
                    item(p.nome, f.rotulo, f.preco)
            else:
                item(p.nome, p.descricao, p.preco)

    # Catálogo vazio: exemplo no formato do cliente, para ele se guiar.
    if not algum:
        exemplos = [
            ("REFEIÇÕES", [
                ("Pequena - 300g", "1 proteína + 1 acompanhamento", Decimal("29.90")),
                ("Média - 500g", "1 proteína + 2 acompanhamentos", Decimal("49.90")),
                ("Grande - 700g", "1 proteína + 3 acompanhamentos", Decimal("69.90")),
            ]),
            ("PORÇÕES", [
                ("Batata Frita", "300g", Decimal("32.00")),
                ("Proteínas", "Média - 500g", Decimal("69.90")),
                ("Proteínas", "Grande - 700g", Decimal("94.90")),
            ]),
            ("SANDUÍCHES", [
                ("Carne Assada", "", Decimal("19.90")),
                ("Frango Desfiado", "", Decimal("14.90")),
                ("Adicionais", "", Decimal("3.00")),
            ]),
            ("BEBIDAS", [
                ("Coca Lata", "", Decimal("9.00")),
                ("Guaraná Lata", "", Decimal("9.00")),
            ]),
            ("SOBREMESA", [("Pudim de Leite", "", Decimal("16.90"))]),
        ]
        for titulo, itens in exemplos:
            secao(titulo)
            for nome, tam, pr in itens:
                item(nome, tam, pr)

    # Aba 2: instruções de como preencher (a aba 1 fica só com os dados).
    ws2 = wb.create_sheet("Como preencher")
    ws2.column_dimensions["A"].width = 95
    instrucoes = [
        "COMO PREENCHER O CARDÁPIO",
        "",
        "1) Nome da SEÇÃO em MAIÚSCULAS, numa linha sozinha (ex.: BEBIDAS).",
        "   Deixe as colunas B e C VAZIAS nessa linha — é assim que ela vira uma categoria.",
        "",
        "2) Abaixo, um item por linha:",
        "      Coluna A = nome do item (ex.: Coca Lata)",
        "      Coluna B = tamanho/descrição — opcional (ex.: 500ml, 300g)",
        "      Coluna C = preço (ex.: R$ 9,00)",
        "",
        "3) Item com VÁRIOS TAMANHOS: repita o nome em linhas seguidas, mudando B e o preço.",
        "      Proteínas | Média - 500g | R$ 69,90",
        "      Proteínas | Grande - 700g | R$ 94,90",
        "   (isso vira um item com opções de tamanho)",
        "",
        "4) Um item chamado 'Adicionais' vira um adicional (extra).",
        "",
        "Pode enviar em Excel (.xlsx) ou CSV (.csv).",
    ]
    for i, txt in enumerate(instrucoes, 1):
        c = ws2.cell(row=i, column=1, value=txt)
        if i == 1:
            c.font = Font(bold=True, size=13, color="B45309")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="modelo_cardapio_bigkilo.xlsx"'
    wb.save(response)
    return response


@require_http_methods(["GET", "POST"])
def importar_planilha_view(request):
    if request.method == "POST":
        arquivo = request.FILES.get("planilha")
        nome_cardapio = request.POST.get("nome_cardapio", "").strip()
        exclusivo = request.POST.get("exclusivo") == "on"

        if not nome_cardapio:
            messages.error(request, "Informe o nome do novo cardápio.")
            return redirect("admin:cardapio_cardapio_changelist")

        if not arquivo:
            messages.error(request, "Nenhum arquivo enviado.")
            return redirect("admin:cardapio_cardapio_changelist")

        nome_arq = (arquivo.name or "").lower()
        if not (nome_arq.endswith(".xlsx") or nome_arq.endswith(".csv")):
            messages.error(request, "O arquivo deve ser Excel (.xlsx) ou CSV (.csv).")
            return redirect("admin:cardapio_cardapio_changelist")

        try:
            rows = _ler_linhas(arquivo)
            if not rows:
                messages.error(request, "A planilha está vazia.")
                return redirect("admin:cardapio_cardapio_importar_planilha")

            template = _e_formato_template(rows)

            # Tudo dentro de UMA transação: se algo falhar no meio, NADA é gravado
            # (evita deixar cardápios órfãos/vazios no banco em caso de erro).
            with transaction.atomic():
                cardapio_obj, tipo_cardapio, qtd_dias = _criar_cardapio_do_form(
                    request, nome_cardapio, exclusivo
                )
                if template:
                    criados, existentes = _importar_template(rows, cardapio_obj)
                else:
                    criados, existentes = _importar_livre(rows, cardapio_obj)

                total = criados + existentes
                if total == 0:
                    raise ValueError(
                        "Nenhum produto válido encontrado na planilha. Confira o conteúdo do arquivo."
                    )

            formato_txt = "modelo padrão" if template else "formato do cliente"
            detalhe = (
                f"{criados} novo(s), {existentes} já existente(s) vinculado(s)."
                if template
                else f"{criados} novo(s), {existentes} atualizado(s)."
            )
            messages.success(
                request,
                f'Cardápio "{nome_cardapio}" criado com {total} produto(s) (lido no {formato_txt})! {detalhe}',
            )
            if not template:
                messages.info(
                    request,
                    "ℹ️ Como a planilha veio no formato do cliente, confira as categorias em "
                    "Categorias (o tipo é adivinhado pelo nome — ajuste se precisar).",
                )
            if tipo_cardapio == "NORMAL" and qtd_dias == 0:
                messages.warning(
                    request,
                    "⚠️ Você não marcou nenhum dia da semana. Assim o cardápio NÃO vai "
                    "aparecer para o cliente. Clique em Editar e escolha os dias/horários.",
                )
            elif tipo_cardapio == "ESPECIAL" and not request.POST.get("data_inicio"):
                messages.warning(
                    request,
                    "⚠️ Cardápio especial sem data inicial: ele não vai aparecer até você definir a data.",
                )
        except Exception as e:
            messages.error(request, f"Erro ao processar a planilha: {str(e)}")

        return redirect("admin:cardapio_cardapio_changelist")

    return render(request, "admin/cardapio/importar_planilha.html")
