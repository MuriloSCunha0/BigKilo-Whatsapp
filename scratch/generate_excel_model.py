import openpyxl
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cardápio"

headers = [
    "Categoria",
    "Tipo da Categoria",
    "Produto",
    "Descrição",
    "Modo de Venda",
    "Preço",
    "Preço por KG",
    "Horário Específico",
    "Ativo",
]

for col_num, nome in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=nome)
    cell.font = openpyxl.styles.Font(bold=True)

rows = [
    ["Sobremesas E2E", "SOBREMESA", "Pudim de Leite Ninho", "Pudim cremoso individual", "UNIDADE", "14.90", "0.00", "", "SIM"],
    ["Bebidas E2E", "BEBIDA", "Suco Natural de Laranja 500ml", "100% fruta sem açúcar", "UNIDADE", "9.50", "0.00", "", "SIM"],
    ["Proteínas E2E", "PROTEINA", "Filé de Frango Grelhado", "Suculento grelhado na hora", "MONTAGEM", "0.00", "0.00", "", "SIM"],
    ["Guarnições E2E", "ACOMPANHAMENTO", "Purê de Batata Baroa", "Cremoso com manteiga", "MONTAGEM", "0.00", "0.00", "", "SIM"],
]

for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

os.makedirs("scratch", exist_ok=True)
target_path = os.path.abspath("scratch/test_menu_import.xlsx")
wb.save(target_path)
print(f"Excel file created at: {target_path}")
